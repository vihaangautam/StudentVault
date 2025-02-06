import openpyxl
import os
import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl.styles import Font

# Define the Excel file path
FILE_NAME = "C:\\Users\\ASUS\\PycharmProjects\\pythonProject1\\StudentDatabaseFINAL.xlsx"

# Function to create a new Excel file with headers
def create_excel_file():
    if not os.path.exists(FILE_NAME):  # Prevent overwriting existing data
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Students"

        # Create a header row with styling
        headers = ["Name", "ID", "Attendance (%)", "Email", "Phone", "Address", "Course", "Year"]
        sheet.append(headers)

        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).font = Font(bold=True)

        workbook.save(FILE_NAME)

# Function to add a student record
def add_student():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    # Get user input via a Tkinter dialog
    name = simpledialog.askstring("Input", "Enter Student's Name:")
    student_id = simpledialog.askinteger("Input", "Enter Student ID:")
    attendance = simpledialog.askinteger("Input", "Enter Attendance Percentage:")
    email = simpledialog.askstring("Input", "Enter Email:")
    phone = simpledialog.askstring("Input", "Enter Phone Number:")
    address = simpledialog.askstring("Input", "Enter Address:")
    course = simpledialog.askstring("Input", "Enter Course:")
    year = simpledialog.askstring("Input", "Enter Year:")

    if not (name and student_id and attendance and email and phone and address and course and year):
        messagebox.showerror("Error", "All fields are required!")
        return

    student_data = [name, student_id, attendance, email, phone, address, course, year]
    sheet.append(student_data)
    workbook.save(FILE_NAME)

    messagebox.showinfo("Success", "Student record added successfully!")

# Function to read and display student data
def read_students():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    students = ""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        students += f"Name: {row[0]}, ID: {row[1]}, Attendance: {row[2]}%\n"
        students += f"Email: {row[3]}, Phone: {row[4]}, Address: {row[5]}\n"
        students += f"Course: {row[6]}, Year: {row[7]}\n\n"

    if students:
        messagebox.showinfo("Student Data", students)
    else:
        messagebox.showinfo("Student Data", "No student records found!")

# Function to search for a student by ID
def search_student():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    search_id = simpledialog.askinteger("Input", "Enter Student ID to Search:")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == search_id:
            student_info = f"Name: {row[0]}, ID: {row[1]}, Attendance: {row[2]}%\n"
            student_info += f"Email: {row[3]}, Phone: {row[4]}, Address: {row[5]}\n"
            student_info += f"Course: {row[6]}, Year: {row[7]}"
            messagebox.showinfo("Student Found", student_info)
            return

    messagebox.showerror("Not Found", f"No student found with ID {search_id}")

# Function to delete a student by ID
def delete_student():
    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook.active

    delete_id = simpledialog.askinteger("Input", "Enter Student ID to Delete:")
    found = False

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=2).value == delete_id:
            sheet.delete_rows(row)
            workbook.save(FILE_NAME)
            messagebox.showinfo("Success", f"Student with ID {delete_id} deleted.")
            found = True
            break

    if not found:
        messagebox.showerror("Not Found", f"No student found with ID {delete_id}")

# Function to display the main menu using Tkinter
def main_menu():
    root = tk.Tk()
    root.title("School Management System")
    root.geometry("400x500")

    tk.Label(root, text="School Management System", font=("Arial", 16, "bold")).pack(pady=10)

    tk.Button(root, text="Add Student", command=add_student, width=25, height=2).pack(pady=5)
    tk.Button(root, text="View Students", command=read_students, width=25, height=2).pack(pady=5)
    tk.Button(root, text="Search Student", command=search_student, width=25, height=2).pack(pady=5)
    tk.Button(root, text="Delete Student", command=delete_student, width=25, height=2).pack(pady=5)
    tk.Button(root, text="Exit", command=root.quit, width=25, height=2, bg="red", fg="white").pack(pady=5)

    root.mainloop()

# Run the program
if __name__ == "__main__":
    create_excel_file()
    main_menu()
